"""
Читання додаткових даних з Dodatky.xlsx
- ЖБД (журнали бойових дій)
- Громади
- Населені пункти
"""
from openpyxl import load_workbook
from datetime import datetime, date
from typing import List, Tuple, Dict, Optional
import os
import re


class DodatkyReader:
    """Клас для читання даних з Dodatky.xlsx"""

    def __init__(self, file_path: str = "Dodatky.xlsx"):
        self.file_path = file_path
        self.zbd_data = []  # [(назва, номер, дата), ...]
        self.hromady_data = []  # [(назва, дата_від, дата_до), ...]
        self.np_data = []  # [(назва, дата_від, дата_до), ...]
        self._loaded = False

    def load(self):
        """Завантажити дані з Excel"""
        if self._loaded:
            return

        if not os.path.exists(self.file_path):
            print(f"[WARNING] Файл {self.file_path} не знайдено")
            return

        wb = load_workbook(self.file_path, read_only=True, data_only=True)

        # Читаємо ЖБД
        if "ЖБД" in wb.sheetnames:
            ws = wb["ЖБД"]
            current_name = None
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0]:  # Якщо є назва ЖБД
                    current_name = str(row[0]).strip()
                if current_name and row[1] and row[2]:  # Номер і дата
                    nomer = str(row[1]).strip()
                    data = self._parse_date(row[2])
                    if data:
                        self.zbd_data.append((current_name, nomer, data))

        # Читаємо Громади (аркуш може називатись "Громада" або "Громади")
        hromady_sheet = None
        for name in ["Громада", "Громади", "ГРОМАДА", "ГРОМАДИ"]:
            if name in wb.sheetnames:
                hromady_sheet = name
                break

        if hromady_sheet:
            ws = wb[hromady_sheet]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] and row[1] and row[2]:
                    nazva = str(row[0]).strip()
                    data_vid = self._parse_date(row[1])
                    data_do = self._parse_date(row[2])
                    if data_vid and data_do:
                        self.hromady_data.append((nazva, data_vid, data_do))

        # Читаємо Населені пункти (різні варіанти назви)
        np_sheet = None
        for name in ["Населений пункт", "Населенні пункт", "Населені пункти", "НП"]:
            if name in wb.sheetnames:
                np_sheet = name
                break

        if np_sheet:
            ws = wb[np_sheet]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] and row[1] and row[2]:
                    nazva = str(row[0]).strip()
                    data_vid = self._parse_date(row[1])
                    data_do = self._parse_date(row[2])
                    if data_vid and data_do:
                        self.np_data.append((nazva, data_vid, data_do))

        wb.close()
        self._loaded = True

    def _parse_date(self, value) -> Optional[date]:
        """Парсить дату з різних форматів"""
        if value is None:
            return None

        if isinstance(value, datetime):
            return value.date()

        if isinstance(value, date):
            return value

        # Спробувати розпарсити рядок
        date_str = str(value).strip()
        formats = ["%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y"]

        for fmt in formats:
            try:
                return datetime.strptime(date_str, fmt).date()
            except ValueError:
                continue

        return None

    def get_zbd(self, periods_text: str) -> str:
        """
        Отримати ЖБД для періодів

        Логіка:
        1. Для кожного періоду знайти всі ЖБД, де дата ЖБД входить в період
        2. Знайти один попередній ЖБД (найближчий до першої дати участі)
        3. Групувати по назві ЖБД, сортувати по датах

        Args:
            periods_text: Текст періодів "з 01.01.2022 по 15.01.2022; з 20.02.2022 по 28.02.2022"

        Returns:
            Текст ЖБД: "ЖБД 1СБ №51/ВП від 11.05.2022, №173/ВП від 05.11.2022"
        """
        self.load()

        if not self.zbd_data or not periods_text:
            return ""

        # Парсимо періоди
        periods = self._parse_periods_text(periods_text)
        if not periods:
            return ""

        # Знаходимо першу дату участі
        first_date = min(p[0] for p in periods)

        # Список для збору всіх ЖБД з повною інформацією (дата, назва, номер)
        zbd_entries = []
        used_keys = set()

        # Знаходимо найближчий попередній ЖБД
        closest_prev = None
        closest_prev_date = None

        for zbd_name, zbd_nomer, zbd_date in self.zbd_data:
            # Перевіряємо чи ЖБД входить в хоча б один період
            for period_start, period_end in periods:
                if period_start <= zbd_date <= period_end:
                    key = f"{zbd_name}||{zbd_date.isoformat()}||{zbd_nomer}"
                    if key not in used_keys:
                        used_keys.add(key)
                        zbd_entries.append((zbd_date, zbd_name, zbd_nomer))
                    break

            # Перевіряємо чи це найближчий попередній
            if zbd_date < first_date:
                if closest_prev_date is None or zbd_date > closest_prev_date:
                    closest_prev_date = zbd_date
                    closest_prev = (zbd_name, zbd_nomer, zbd_date)

        # Додаємо один попередній ЖБД
        if closest_prev:
            zbd_name, zbd_nomer, zbd_date = closest_prev
            key = f"{zbd_name}||{zbd_date.isoformat()}||{zbd_nomer}"
            if key not in used_keys:
                zbd_entries.append((zbd_date, zbd_name, zbd_nomer))

        # Сортуємо по даті
        zbd_entries.sort(key=lambda x: x[0])

        # Формуємо результат - повторюємо назву для кожного номера
        # Формат: "ЖБД 1СБ №51/ВП від 11.05.2022, ЖБД 1СБ №173/ВП від 05.11.2022"
        result_parts = []
        for zbd_date, zbd_name, zbd_nomer in zbd_entries:
            result_parts.append(f"{zbd_name} №{zbd_nomer} від {zbd_date.strftime('%d.%m.%Y')}")

        return ", ".join(result_parts)

    def get_hromady(self, periods_text: str) -> str:
        """
        Отримати громади для періодів

        Логіка:
        Для кожного періоду перевірити чи він перетинається з періодами громад
        Сортувати громади в хронологічному порядку (за датою початку)

        Args:
            periods_text: Текст періодів

        Returns:
            Текст громад: "Громада1, Громада2" (в хронологічному порядку)
        """
        self.load()

        if not self.hromady_data or not periods_text:
            return ""

        periods = self._parse_periods_text(periods_text)
        if not periods:
            return ""

        # Збираємо громади з датами (дата_від, назва)
        found = []
        found_names = set()  # Щоб уникнути дублікатів

        for period_start, period_end in periods:
            for hromada_name, hromada_vid, hromada_do in self.hromady_data:
                # Перевіряємо перетин періодів
                if period_end >= hromada_vid and period_start <= hromada_do:
                    if hromada_name not in found_names:
                        found_names.add(hromada_name)
                        found.append((hromada_vid, hromada_name))

        # Сортуємо по даті початку (хронологічно)
        found.sort(key=lambda x: x[0])

        # Повертаємо тільки назви, без дат
        return ", ".join([name for _, name in found])

    def get_np(self, periods_text: str) -> str:
        """
        Отримати населені пункти для періодів

        Args:
            periods_text: Текст періодів

        Returns:
            Текст НП: "НП1, НП2"
        """
        self.load()

        if not self.np_data or not periods_text:
            return ""

        periods = self._parse_periods_text(periods_text)
        if not periods:
            return ""

        found = set()

        for period_start, period_end in periods:
            for np_name, np_vid, np_do in self.np_data:
                # Перевіряємо перетин періодів
                if period_end >= np_vid and period_start <= np_do:
                    found.add(np_name)

        return ", ".join(sorted(found))

    def _parse_periods_text(self, periods_text: str) -> List[Tuple[date, date]]:
        """Парсить текст періодів в список кортежів (дата_початку, дата_кінця)"""
        periods = []

        # Розбиваємо по ;
        parts = periods_text.split(";")

        for part in parts:
            part = part.strip()
            if "з " in part and " по " in part:
                # Витягуємо дати
                match = re.search(r'з\s+(\d{2}\.\d{2}\.\d{4})\s+по\s+(\d{2}\.\d{2}\.\d{4})', part)
                if match:
                    try:
                        start = datetime.strptime(match.group(1), "%d.%m.%Y").date()
                        end = datetime.strptime(match.group(2), "%d.%m.%Y").date()
                        periods.append((start, end))
                    except ValueError:
                        continue

        return periods


# Глобальний екземпляр для кешування
_dodatky_reader = None

def get_dodatky_reader(file_path: str = None) -> DodatkyReader:
    """Отримати глобальний екземпляр DodatkyReader"""
    global _dodatky_reader
    if _dodatky_reader is None:
        # Якщо шлях не вказано, шукаємо файл відносно виконуваного файлу
        if file_path is None:
            import os
            import sys

            # Визначаємо базову директорію (для exe та для звичайного Python)
            if getattr(sys, 'frozen', False):
                # Запущено як exe (PyInstaller)
                base_dir = sys._MEIPASS
            else:
                # Запущено як Python скрипт
                base_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

            file_path = os.path.join(base_dir, "Dodatky.xlsx")

        _dodatky_reader = DodatkyReader(file_path)
    return _dodatky_reader
