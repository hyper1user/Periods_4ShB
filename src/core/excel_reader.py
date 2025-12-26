"""
Модуль для читання та парсингу Excel файлу з даними військовослужбовців
"""
from openpyxl import load_workbook
from typing import Optional
import os


class ExcelReader:
    """
    Клас для читання даних з Excel файлу
    """

    def __init__(self, file_path: str):
        """
        Ініціалізація читача Excel

        Args:
            file_path: Шлях до Excel файлу
        """
        self.file_path = file_path
        self.workbook = None

    def load_workbook(self):
        """
        Завантаження Excel файлу

        Raises:
            FileNotFoundError: Якщо файл не знайдено
            Exception: Якщо помилка при відкритті файлу
        """
        if not os.path.exists(self.file_path):
            raise FileNotFoundError(f"Файл не знайдено: {self.file_path}")

        try:
            self.workbook = load_workbook(
                self.file_path,
                read_only=False,  # Дозволяємо запис для додавання даних
                data_only=True,   # Читати значення, а не формули
                keep_vba=False    # Не завантажувати VBA макроси - ШВИДШЕ
            )
        except Exception as e:
            raise Exception(f"Помилка при відкритті файлу: {str(e)}")

    def get_sheet_data(self, sheet_name: str) -> list[dict]:
        """
        Отримати всі дані з конкретного аркуша

        Args:
            sheet_name: Назва аркуша

        Returns:
            Список словників з даними

        Raises:
            ValueError: Якщо аркуш не знайдено
        """
        if not self.workbook:
            self.load_workbook()

        if sheet_name not in self.workbook.sheetnames:
            raise ValueError(f"Аркуш '{sheet_name}' не знайдено")

        sheet = self.workbook[sheet_name]
        data = []

        # Залежно від аркуша, читаємо різні стовпці
        if sheet_name == "Data":
            # Стовпці: A=місяць, B=підрозділ, D=звання, E=ПІБ, F=РНОКПП, G=посада
            # H=початок 100%, I=кінець 100%, J=початок 30%, K=кінець 30%
            # L=початок не залучення, M=кінець не залучення, N=статус
            # AB=дата народження (індекс 27)
            for row_idx, row in enumerate(sheet.iter_rows(min_row=4, values_only=True), start=4):
                if row[4]:  # Якщо є ПІБ (стовпець E, індекс 4)
                    # Форматуємо дату народження зі стовпця AB (індекс 27)
                    birth_date = None
                    if len(row) > 27 and row[27]:
                        from datetime import datetime
                        if isinstance(row[27], datetime):
                            birth_date = row[27].strftime("%d.%m.%Y")
                        else:
                            birth_date = str(row[27])

                    data.append({
                        "month": str(row[0]) if row[0] else None,
                        "unit": str(row[1]) if row[1] else None,
                        "birth_date": birth_date,
                        "rank": str(row[3]) if row[3] else None,
                        "name": str(row[4]) if row[4] else None,
                        "rnokpp": str(row[5]) if row[5] else None,
                        "position": str(row[6]) if row[6] else None,
                        "start_100": row[7],
                        "end_100": row[8],
                        "start_30": row[9],
                        "end_30": row[10],
                        "start_non": row[11],
                        "end_non": row[12],
                        "status": str(row[13]) if len(row) > 13 and row[13] else None,
                        "row_number": row_idx  # НОВЕ: для синхронізації
                    })

        elif sheet_name in ["Періоди на 100", "Періоди на 30", "Періоди не залучення"]:
            # Стовпці: A=№ п/п або місяць, B=ПІБ, C=періоди (текст)
            # ВАЖЛИВО: Дані починаються з рядка 2, рядок 1 - заголовки
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if row[1]:  # Якщо є ПІБ (стовпець B, індекс 1)
                    data.append({
                        "month": str(row[0]) if row[0] else None,
                        "name": str(row[1]) if row[1] else None,
                        "periods": str(row[2]) if len(row) > 2 and row[2] else None,
                        "row_number": row_idx  # НОВЕ: для синхронізації
                    })

        return data

    def get_unique_names(self, sheet_name: str = "Data") -> list[str]:
        """
        Отримати список унікальних ПІБ

        Args:
            sheet_name: Назва аркуша (за замовчуванням "Data")

        Returns:
            Відсортований список унікальних ПІБ
        """
        data = self.get_sheet_data(sheet_name)
        names = set()

        for row in data:
            if row.get("name"):
                names.add(row["name"])

        return sorted(list(names))

    def get_unique_units(self) -> list[str]:
        """
        Отримати список унікальних підрозділів з аркуша Data

        Returns:
            Відсортований список унікальних підрозділів
        """
        data = self.get_sheet_data("Data")
        units = set()

        for row in data:
            if row.get("unit"):
                units.add(row["unit"])

        return sorted(list(units))

    def get_servicemember_data(self, name: str, sheet_name: str) -> list[dict]:
        """
        Отримати всі дані для конкретного військовослужбовця з аркуша

        Args:
            name: ПІБ військовослужбовця
            sheet_name: Назва аркуша

        Returns:
            Список словників з даними для військовослужбовця
        """
        data = self.get_sheet_data(sheet_name)
        return [row for row in data if row.get("name") == name]

    def get_unit_data(self, unit: str, sheet_name: str) -> list[dict]:
        """
        Отримати всі дані для конкретного підрозділу

        Args:
            unit: Назва підрозділу
            sheet_name: Назва аркуша

        Returns:
            Список словників з даними для підрозділу
        """
        data = self.get_sheet_data(sheet_name)
        return [row for row in data if row.get("unit") == unit]

    def get_servicemember_info_from_data(self, name: str) -> Optional[dict]:
        """
        Отримати основну інформацію про військовослужбовця з аркуша Data
        Повертає останній запис (найновіші звання та посада)
        Для РНОКПП та дати народження шукає у всіх записах

        Args:
            name: ПІБ військовослужбовця

        Returns:
            Словник з інформацією або None
        """
        data = self.get_sheet_data("Data")

        # Знаходимо всі записи для цього військовослужбовця
        matching_rows = [row for row in data if row.get("name") == name]

        if not matching_rows:
            return None

        # Повертаємо останній запис (найновіші звання та посада)
        last_row = matching_rows[-1]

        # Шукаємо РНОКПП у всіх записах (беремо перше непорожнє значення)
        rnokpp = last_row.get("rnokpp", "")
        if not rnokpp or rnokpp == "None" or not str(rnokpp).strip():
            for row in matching_rows:
                temp_rnokpp = row.get("rnokpp", "")
                if temp_rnokpp and temp_rnokpp != "None" and str(temp_rnokpp).strip():
                    rnokpp = temp_rnokpp
                    break

        # Шукаємо дату народження у всіх записах (беремо перше непорожнє значення)
        birth_date = last_row.get("birth_date")
        if not birth_date or birth_date == "None" or not str(birth_date).strip():
            for row in matching_rows:
                temp_birth = row.get("birth_date")
                if temp_birth and temp_birth != "None" and str(temp_birth).strip():
                    birth_date = temp_birth
                    break

        # Перша літера посади - маленька
        position = last_row["position"]
        if position and len(position) > 0:
            position = position[0].lower() + position[1:]

        return {
            "name": last_row["name"],
            "rank": last_row["rank"],
            "position": position,
            "rnokpp": rnokpp if rnokpp and str(rnokpp) != "None" else "",
            "unit": last_row["unit"],
            "birth_date": birth_date if birth_date and str(birth_date) != "None" else ""
        }

    def add_servicemember_data(self, data: dict) -> bool:
        """
        Додає нові дані військовослужбовця в аркуш Data

        Args:
            data: Словник з даними {
                "month": "місяць",
                "unit": "підрозділ",
                "birth_date": "дата народження",
                "rank": "звання",
                "name": "ПІБ",
                "rnokpp": "РНОКПП",
                "position": "посада"
            }

        Returns:
            True якщо успішно, False якщо помилка
        """
        try:
            sheet = self.workbook["Data"]

            # Знаходимо наступний вільний рядок
            next_row = sheet.max_row + 1

            # Додаємо дані (стовпці A-G, відповідно до структури)
            sheet.cell(row=next_row, column=1, value=data.get("month", ""))  # A - місяць
            sheet.cell(row=next_row, column=2, value=data.get("unit", ""))   # B - підрозділ
            sheet.cell(row=next_row, column=28, value=data.get("birth_date", ""))  # AB - дата народження
            sheet.cell(row=next_row, column=4, value=data.get("rank", ""))   # D - звання
            sheet.cell(row=next_row, column=5, value=data.get("name", ""))   # E - ПІБ
            sheet.cell(row=next_row, column=6, value=data.get("rnokpp", "")) # F - РНОКПП
            sheet.cell(row=next_row, column=7, value=data.get("position", ""))  # G - посада

            return True
        except Exception as e:
            print(f"Помилка при додаванні даних: {str(e)}")
            return False

    def add_period(self, name: str, sheet_name: str, period: str) -> bool:
        """
        Додає період для військовослужбовця

        Args:
            name: ПІБ військовослужбовця
            sheet_name: Назва аркуша ("Періоди на 100" або "Періоди на 30")
            period: Текст періоду (наприклад "з 01.12.2024 по 31.12.2024")

        Returns:
            True якщо успішно, False якщо помилка
        """
        try:
            if sheet_name not in self.workbook.sheetnames:
                print(f"Аркуш {sheet_name} не знайдено")
                return False

            sheet = self.workbook[sheet_name]

            # Знаходимо останній рядок з даними (де стовпець B має значення)
            # Рядок 1 - заголовки, дані починаються з рядка 2
            last_data_row = 1  # Рядок заголовків

            # Шукаємо з кінця для оптимізації
            for row in range(sheet.max_row, 1, -1):
                cell_value = sheet.cell(row=row, column=2).value  # Стовпець B (ПІБ)
                if cell_value and str(cell_value).strip() and str(cell_value) != "None":
                    last_data_row = row
                    break

            next_row = last_data_row + 1

            # Додаємо дані (A=місяць, B=ПІБ, C=періоди)
            sheet.cell(row=next_row, column=1, value="")  # A - місяць (опціонально)
            sheet.cell(row=next_row, column=2, value=name)  # B - ПІБ
            sheet.cell(row=next_row, column=3, value=period)  # C - періоди

            return True
        except Exception as e:
            print(f"Помилка при додаванні періоду: {str(e)}")
            return False

    def save(self) -> bool:
        """
        Зберігає зміни у файл

        Returns:
            True якщо успішно, False якщо помилка
        """
        try:
            self.workbook.save(self.file_path)
            return True
        except Exception as e:
            print(f"Помилка при збереженні файлу: {str(e)}")
            return False

    def close(self):
        """
        Закрити workbook
        """
        if self.workbook:
            self.workbook.close()
            self.workbook = None
