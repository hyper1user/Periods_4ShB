"""
Діалог для імпорту даних з іншого Excel файлу
"""
from PySide6.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QLabel,
    QLineEdit, QPushButton, QComboBox, QMessageBox,
    QFileDialog, QGroupBox, QProgressDialog, QStackedWidget, QCheckBox, QWidget
)
from PySide6.QtCore import Qt
from openpyxl import load_workbook
from datetime import datetime
import os
from utils.paths import get_base_dir


class ImportDataDialog(QDialog):
    """
    Діалог-wizard для імпорту даних за місяць з іншого Excel файлу
    АДАПТОВАНО: працює з ExcelReader або DatabaseManager
    """

    def __init__(self, data_source, use_database=False, parent=None):
        super().__init__(parent)
        self.data_source = data_source  # ExcelReader або DatabaseManager
        self.use_database = use_database  # Чи використовувати БД

        # Backwards compatibility
        self.excel_reader = data_source if not use_database else None
        self.db_manager = data_source if use_database else None

        # Дані для кожного кроку
        self.step_data = {
            "100": {
                "enabled": False,
                "workbook": None,
                "file_path": "",
                "sheet_name": "",
                "rank_col": "",
                "name_col": "",
                "position_col": "",
                "start_col": "",
                "end_col": ""
            },
            "30": {
                "enabled": False,
                "workbook": None,
                "file_path": "",
                "sheet_name": "",
                "rank_col": "",
                "name_col": "",
                "position_col": "",
                "start_col": "",
                "end_col": ""
            },
            "non": {
                "enabled": False,
                "workbook": None,
                "file_path": "",
                "sheet_name": "",
                "rank_col": "",
                "name_col": "",
                "position_col": "",
                "start_col": "",
                "end_col": ""
            }
        }

        self.current_step = 0
        self.month = ""

        self.init_ui()

    def init_ui(self):
        """
        Ініціалізація інтерфейсу
        """
        self.setWindowTitle("Додати новий місяць")
        self.setMinimumWidth(700)
        self.setMinimumHeight(600)

        layout = QVBoxLayout()

        # Заголовок
        self.title = QLabel("Імпорт даних за місяць")
        self.title.setStyleSheet("font-size: 16px; font-weight: bold; color: #1976d2;")
        layout.addWidget(self.title)

        # Індикатор кроку
        self.step_indicator = QLabel("Крок 1 з 4")
        self.step_indicator.setStyleSheet("font-size: 12px; color: #666;")
        layout.addWidget(self.step_indicator)

        layout.addSpacing(10)

        # Stacked widget для кроків
        self.stacked_widget = QStackedWidget()

        # Створюємо всі кроки
        self.stacked_widget.addWidget(self.create_month_step())  # 0
        self.stacked_widget.addWidget(self.create_period_step("100", "Періоди на 100 тис.", "#2e7d32"))  # 1
        self.stacked_widget.addWidget(self.create_period_step("30", "Періоди на 30 тис.", "#ff6f00"))  # 2
        self.stacked_widget.addWidget(self.create_period_step("non", "Періоди не залучення", "#c62828"))  # 3

        layout.addWidget(self.stacked_widget)

        # Кнопки навігації
        button_layout = QHBoxLayout()

        self.btn_back = QPushButton("← Назад")
        self.btn_back.clicked.connect(self.go_back)
        self.btn_back.setEnabled(False)
        button_layout.addWidget(self.btn_back)

        button_layout.addStretch()

        self.btn_next = QPushButton("Далі →")
        self.btn_next.clicked.connect(self.go_next)
        self.btn_next.setStyleSheet("background-color: #1976d2; color: white; font-weight: bold; padding: 10px;")
        button_layout.addWidget(self.btn_next)

        self.btn_finish = QPushButton("Завершити імпорт")
        self.btn_finish.clicked.connect(self.import_data)
        self.btn_finish.setStyleSheet("background-color: #4caf50; color: white; font-weight: bold; padding: 10px;")
        self.btn_finish.setVisible(False)
        button_layout.addWidget(self.btn_finish)

        btn_cancel = QPushButton("Скасувати")
        btn_cancel.clicked.connect(self.reject)
        button_layout.addWidget(btn_cancel)

        layout.addLayout(button_layout)
        self.setLayout(layout)

    def create_month_step(self):
        """
        Крок 0: Вказання місяця
        """
        widget = QGroupBox("Вкажіть місяць для імпорту")
        layout = QVBoxLayout()

        info = QLabel("Введіть місяць у форматі YYYY-MM (наприклад: 2025-08)")
        info.setStyleSheet("color: #666; font-style: italic;")
        layout.addWidget(info)

        layout.addSpacing(20)

        month_layout = QHBoxLayout()
        month_layout.addWidget(QLabel("Місяць:"))
        self.month_input = QLineEdit()
        self.month_input.setPlaceholderText("2025-11")
        self.month_input.setStyleSheet("QLineEdit { color: black; background-color: white; }")
        month_layout.addWidget(self.month_input)
        layout.addLayout(month_layout)

        layout.addStretch()
        widget.setLayout(layout)
        return widget

    def create_period_step(self, step_key, title, color):
        """
        Створює крок для імпорту конкретного типу періодів
        """
        widget = QWidget()
        layout = QVBoxLayout()

        # Заголовок кроку
        step_title = QLabel(title)
        step_title.setStyleSheet(f"font-size: 14px; font-weight: bold; color: {color};")
        layout.addWidget(step_title)

        # Чекбокс для пропуску кроку
        checkbox = QCheckBox(f"Імпортувати {title.lower()}")
        checkbox.setChecked(False)
        checkbox.stateChanged.connect(lambda state, key=step_key: self.toggle_step(key, state))
        layout.addWidget(checkbox)
        setattr(self, f"checkbox_{step_key}", checkbox)

        layout.addSpacing(10)

        # Контейнер для полів (спочатку вимкнений)
        fields_container = QGroupBox("Налаштування імпорту")
        fields_layout = QVBoxLayout()

        # Вибір файлу
        file_group = QLabel("1. Оберіть файл джерела:")
        file_group.setStyleSheet("font-weight: bold;")
        fields_layout.addWidget(file_group)

        file_layout = QHBoxLayout()
        file_label = QLabel("Файл не обрано")
        file_label.setStyleSheet("color: gray;")
        file_layout.addWidget(file_label)
        setattr(self, f"file_label_{step_key}", file_label)

        file_btn = QPushButton("Обрати файл...")
        file_btn.clicked.connect(lambda checked, key=step_key: self.select_file(key))
        file_layout.addWidget(file_btn)
        fields_layout.addLayout(file_layout)

        # Стиль для комбобоксів
        combo_style = """
            QComboBox {
                color: black;
                background-color: white;
            }
            QComboBox:disabled {
                color: gray;
                background-color: #f0f0f0;
            }
            QComboBox::drop-down {
                border: none;
            }
            QComboBox QAbstractItemView {
                color: black;
                background-color: white;
            }
        """

        # Вибір аркушу
        sheet_layout = QHBoxLayout()
        sheet_layout.addWidget(QLabel("Аркуш:"))
        sheet_combo = QComboBox()
        sheet_combo.setStyleSheet(combo_style)
        sheet_combo.setEnabled(False)
        sheet_combo.currentTextChanged.connect(lambda text, key=step_key: self.on_sheet_changed(key))
        sheet_layout.addWidget(sheet_combo)
        fields_layout.addLayout(sheet_layout)
        setattr(self, f"sheet_combo_{step_key}", sheet_combo)

        fields_layout.addSpacing(15)

        # Мапінг стовпців (випадаючі списки з заголовками)
        mapping_group = QLabel("2. Оберіть колонки:")
        mapping_group.setStyleSheet("font-weight: bold;")
        fields_layout.addWidget(mapping_group)

        unit_combo = self.create_column_combo(fields_layout, "Підрозділ:", combo_style)
        rank_combo = self.create_column_combo(fields_layout, "Звання:", combo_style)
        name_combo = self.create_column_combo(fields_layout, "ПІБ:", combo_style)
        position_combo = self.create_column_combo(fields_layout, "Посада:", combo_style)
        start_combo = self.create_column_combo(fields_layout, "Початок періоду:", combo_style)
        end_combo = self.create_column_combo(fields_layout, "Кінець періоду:", combo_style)

        setattr(self, f"unit_combo_{step_key}", unit_combo)
        setattr(self, f"rank_combo_{step_key}", rank_combo)
        setattr(self, f"name_combo_{step_key}", name_combo)
        setattr(self, f"position_combo_{step_key}", position_combo)
        setattr(self, f"start_combo_{step_key}", start_combo)
        setattr(self, f"end_combo_{step_key}", end_combo)

        fields_layout.addStretch()
        fields_container.setLayout(fields_layout)
        fields_container.setEnabled(False)
        setattr(self, f"fields_container_{step_key}", fields_container)

        layout.addWidget(fields_container)
        widget.setLayout(layout)
        return widget

    def create_column_combo(self, layout, label_text, combo_style):
        """
        Створює випадаючий список для вибору стовпця
        """
        field_layout = QHBoxLayout()
        field_layout.addWidget(QLabel(label_text))
        combo = QComboBox()
        combo.setMinimumWidth(200)
        combo.setStyleSheet(combo_style)
        field_layout.addWidget(combo)
        field_layout.addStretch()
        layout.addLayout(field_layout)
        return combo

    def on_sheet_changed(self, step_key):
        """
        Обробник зміни аркуша - заповнює списки колонок
        """
        workbook = self.step_data[step_key]["workbook"]
        if not workbook:
            return

        sheet_combo = getattr(self, f"sheet_combo_{step_key}")
        sheet_name = sheet_combo.currentText()
        if not sheet_name:
            return

        try:
            sheet = workbook[sheet_name]

            # Отримуємо заголовки з першого рядка
            headers = []
            first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), [])
            for col_idx, cell in enumerate(first_row, 1):
                col_letter = self.index_to_column_letter(col_idx)
                header_text = str(cell) if cell else f"(Колонка {col_letter})"
                headers.append((col_idx, f"{col_letter}: {header_text}"))

            # Отримуємо всі комбобокси для цього кроку
            combos = [
                getattr(self, f"unit_combo_{step_key}"),
                getattr(self, f"rank_combo_{step_key}"),
                getattr(self, f"name_combo_{step_key}"),
                getattr(self, f"position_combo_{step_key}"),
                getattr(self, f"start_combo_{step_key}"),
                getattr(self, f"end_combo_{step_key}"),
            ]

            # Заповнюємо комбобокси
            for combo in combos:
                combo.clear()
                combo.addItem("(не обрано)", None)
                for col_idx, header in headers:
                    combo.addItem(header, col_idx)

            # Автоматично вибираємо колонки за ключовими словами
            self.auto_select_columns(step_key, headers)

        except Exception as e:
            print(f"Помилка при зміні аркуша: {e}")

    def index_to_column_letter(self, col_idx: int) -> str:
        """Конвертує індекс колонки в букву (1=A, 2=B, ...)"""
        result = ""
        while col_idx > 0:
            col_idx, remainder = divmod(col_idx - 1, 26)
            result = chr(65 + remainder) + result
        return result

    def auto_select_columns(self, step_key, headers):
        """Автоматично вибирає колонки за назвами заголовків"""
        keywords = {
            "unit": ["підрозділ", "unit", "підр"],
            "rank": ["звання", "rank", "зван"],
            "name": ["піб", "прізвище", "ім'я", "name", "пiб"],
            "position": ["посада", "position", "посад"],
            "start": ["початок", "start", "з ", "від"],
            "end": ["кінець", "end", "по ", "до"],
        }

        for field_name, field_keywords in keywords.items():
            combo = getattr(self, f"{field_name}_combo_{step_key}")

            for i, (col_idx, header) in enumerate(headers):
                header_lower = header.lower()
                for keyword in field_keywords:
                    if keyword in header_lower:
                        combo.setCurrentIndex(i + 1)  # +1 бо перший елемент "(не обрано)"
                        break

    def toggle_step(self, step_key, state):
        """
        Увімкнення/вимкнення кроку
        """
        enabled = (state == Qt.CheckState.Checked.value)
        self.step_data[step_key]["enabled"] = enabled

        fields_container = getattr(self, f"fields_container_{step_key}")
        fields_container.setEnabled(enabled)

    def select_file(self, step_key):
        """
        Вибір файлу для конкретного кроку
        """
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Оберіть файл джерела",
            "",
            "Excel Files (*.xlsx *.xlsm)"
        )

        if file_path:
            try:
                # Спробуємо різні параметри для проблемних файлів
                workbook = None
                try:
                    workbook = load_workbook(file_path, data_only=True)
                except Exception:
                    try:
                        workbook = load_workbook(file_path, read_only=True, data_only=True)
                    except Exception:
                        workbook = load_workbook(file_path, data_only=False)

                self.step_data[step_key]["workbook"] = workbook
                self.step_data[step_key]["file_path"] = file_path

                file_label = getattr(self, f"file_label_{step_key}")
                file_label.setText(file_path)
                file_label.setStyleSheet("color: green;")

                # Заповнюємо список аркушів
                sheet_combo = getattr(self, f"sheet_combo_{step_key}")
                sheet_combo.clear()
                sheet_combo.addItems(workbook.sheetnames)
                sheet_combo.setEnabled(True)

            except Exception as e:
                error_msg = "Не вдалось відкрити файл!\n\n"
                error_msg += "Можливі причини:\n"
                error_msg += "• Файл пошкоджений або містить невалідні дані\n"
                error_msg += "• Файл .xlsm містить макроси, які конфліктують\n\n"
                error_msg += "Рекомендації:\n"
                error_msg += "• Спробуйте зберегти файл як .xlsx (без макросів)\n"
                error_msg += "• Відкрийте файл в Excel і збережіть знову\n\n"
                error_msg += f"Технічна помилка: {str(e)}"
                QMessageBox.critical(self, "Помилка", error_msg)

    def go_next(self):
        """
        Перехід до наступного кроку
        """
        # Валідація поточного кроку
        if self.current_step == 0:
            month = self.month_input.text().strip()
            if not month:
                QMessageBox.warning(self, "Помилка", "Вкажіть місяць!")
                return
            self.month = month

        self.current_step += 1
        self.update_ui()

    def go_back(self):
        """
        Повернення до попереднього кроку
        """
        self.current_step -= 1
        self.update_ui()

    def update_ui(self):
        """
        Оновлення UI в залежності від поточного кроку
        """
        self.stacked_widget.setCurrentIndex(self.current_step)

        # Оновлення індикатора
        self.step_indicator.setText(f"Крок {self.current_step + 1} з 4")

        # Оновлення кнопок
        self.btn_back.setEnabled(self.current_step > 0)

        if self.current_step == 3:  # Останній крок
            self.btn_next.setVisible(False)
            self.btn_finish.setVisible(True)
        else:
            self.btn_next.setVisible(True)
            self.btn_finish.setVisible(False)

    def column_letter_to_index(self, letter):
        """
        Конвертує букву стовпця в індекс (A=1, B=2, ..., Z=26, AA=27, ...)
        """
        letter = letter.upper().strip()
        result = 0
        for char in letter:
            result = result * 26 + (ord(char) - ord('A') + 1)
        return result

    def import_data(self):
        """
        Імпортує дані з усіх активних кроків
        """
        # Перевірка чи хоча б один крок активний
        if not any(self.step_data[key]["enabled"] for key in ["100", "30", "non"]):
            QMessageBox.warning(self, "Помилка", "Оберіть хоча б один тип періодів для імпорту!")
            return

        # Збираємо дані з кожного кроку
        # ЗМІНА: all_data тепер список записів, а не словник по іменах
        # Це дозволяє мати кілька записів для однієї людини з різними періодами
        all_records = []

        for step_key, step_label in [("100", "100 тис."), ("30", "30 тис."), ("non", "не залучення")]:
            if not self.step_data[step_key]["enabled"]:
                continue

            # Отримуємо дані для цього кроку
            workbook = self.step_data[step_key]["workbook"]
            sheet_combo = getattr(self, f"sheet_combo_{step_key}")
            sheet_name = sheet_combo.currentText()

            if not workbook or not sheet_name:
                QMessageBox.warning(self, "Помилка", f"Не обрано файл або аркуш для '{step_label}'!")
                return

            sheet = workbook[sheet_name]

            # Отримуємо мапінг стовпців з комбобоксів
            try:
                unit_combo = getattr(self, f"unit_combo_{step_key}")
                rank_combo = getattr(self, f"rank_combo_{step_key}")
                name_combo = getattr(self, f"name_combo_{step_key}")
                position_combo = getattr(self, f"position_combo_{step_key}")
                start_combo = getattr(self, f"start_combo_{step_key}")
                end_combo = getattr(self, f"end_combo_{step_key}")

                unit_col = unit_combo.currentData()
                rank_col = rank_combo.currentData()
                name_col = name_combo.currentData()
                position_col = position_combo.currentData()
                start_col = start_combo.currentData()
                end_col = end_combo.currentData()

                # Перевіряємо обов'язкові колонки
                if not name_col:
                    QMessageBox.warning(self, "Помилка", f"Оберіть колонку ПІБ для '{step_label}'!")
                    return
                if not start_col or not end_col:
                    QMessageBox.warning(self, "Помилка", f"Оберіть колонки періоду для '{step_label}'!")
                    return

            except Exception as e:
                QMessageBox.critical(self, "Помилка", f"Помилка при отриманні колонок для '{step_label}': {str(e)}")
                return

            # Читаємо дані з файлу (індекси вже 1-based з currentData)
            for row in sheet.iter_rows(min_row=2, values_only=True):
                name = str(row[name_col - 1]) if name_col <= len(row) and row[name_col - 1] else None
                if not name or name == "None":
                    continue

                unit = str(row[unit_col - 1]) if unit_col and unit_col <= len(row) and row[unit_col - 1] else ""
                rank = str(row[rank_col - 1]) if rank_col and rank_col <= len(row) and row[rank_col - 1] else ""
                position = str(row[position_col - 1]) if position_col and position_col <= len(row) and row[position_col - 1] else ""

                # Отримуємо значення періодів
                start_val = row[start_col - 1] if start_col <= len(row) else None
                end_val = row[end_col - 1] if end_col <= len(row) else None

                if not start_val or not end_val:
                    continue

                # НОВА ЛОГІКА: Парсимо кілька періодів з однієї клітинки
                # Формат: в клітинці може бути кілька рядків (Alt+Enter в Excel):
                #   Початок:     Кінець:
                #   01.08.2025   10.08.2025
                #   17.08.2025   31.08.2025
                # ВАЖЛИВО: openpyxl може повернути це як "01.08.2025\n17.08.2025"
                #          або як "01.08.2025 17.08.2025" (через пробіл!)

                start_str = str(start_val).strip()
                end_str = str(end_val).strip()

                # Розбиваємо по \n (переноси рядків) або по пробілах
                # Якщо є \n - використовуємо його, інакше - пробіли
                if '\n' in start_str:
                    start_dates = start_str.split('\n')
                else:
                    start_dates = start_str.split()

                if '\n' in end_str:
                    end_dates = end_str.split('\n')
                else:
                    end_dates = end_str.split()

                # Для кожної пари (початок, кінець) створюємо окремий запис
                for start_date_str, end_date_str in zip(start_dates, end_dates):
                    start_date_str = start_date_str.strip()
                    end_date_str = end_date_str.strip()

                    if not start_date_str or not end_date_str:
                        continue

                    record = {
                        "name": name,
                        "unit": unit,
                        "rank": rank,
                        "position": position,
                        "start_100": None,
                        "end_100": None,
                        "start_30": None,
                        "end_30": None,
                        "start_non": None,
                        "end_non": None
                    }

                    if step_key == "100":
                        record["start_100"] = self.format_date(start_date_str)
                        record["end_100"] = self.format_date(end_date_str)
                    elif step_key == "30":
                        record["start_30"] = self.format_date(start_date_str)
                        record["end_30"] = self.format_date(end_date_str)
                    elif step_key == "non":
                        record["start_non"] = self.format_date(start_date_str)
                        record["end_non"] = self.format_date(end_date_str)

                    all_records.append(record)

        if not all_records:
            QMessageBox.warning(self, "Помилка", "Не знайдено даних для імпорту!")
            return

        # Підрахунок унікальних осіб
        unique_names = set(r["name"] for r in all_records)

        # ПЕРЕВІРКА НА ДУБЛІКАТИ: Чи вже є записи за цей місяць?
        if self.use_database:
            cursor = self.db_manager.connection.cursor()
            cursor.execute("""
                SELECT COUNT(*) FROM service_records WHERE month = ?
            """, (self.month,))
            existing_count = cursor.fetchone()[0]

            if existing_count > 0:
                reply = QMessageBox.warning(
                    self,
                    "Увага! Дублікати!",
                    f"За місяць {self.month} вже є {existing_count} записів в базі даних!\n\n"
                    f"Імпортування додасть ще {len(all_records)} записів, що створить дублікати.\n\n"
                    f"Продовжити імпорт?",
                    QMessageBox.Yes | QMessageBox.No,
                    QMessageBox.No
                )

                if reply != QMessageBox.Yes:
                    return

        # Підтвердження
        reply = QMessageBox.question(
            self,
            "Підтвердження",
            f"Імпортувати {len(all_records)} записів для {len(unique_names)} людей за місяць {self.month}?",
            QMessageBox.Yes | QMessageBox.No
        )

        if reply != QMessageBox.Yes:
            return

        # Імпорт даних
        try:
            if self.use_database:
                # НОВИЙ ШЛЯХ: Імпорт в БД з прогресом
                # Розраховуємо максимум: кількість записів + перерахунок періодів
                max_progress = len(all_records) * 2  # Приблизно

                progress = QProgressDialog("Імпорт даних в базу даних...", None, 0, max_progress, self)
                progress.setWindowModality(Qt.WindowModal)
                progress.setWindowTitle("Імпорт даних")
                progress.show()

                # Callback для оновлення прогресу
                def update_progress(current, total, message):
                    progress.setLabelText(message)
                    progress.setValue(current)
                    from PySide6.QtWidgets import QApplication
                    QApplication.processEvents()

                # Викликаємо імпорт з callback (передаємо список записів)
                stats = self.db_manager.import_month_data(self.month, all_records, progress_callback=update_progress)

                progress.setValue(max_progress)

                progress.close()

                # Показуємо результат
                message = f"Дані успішно імпортовано в базу даних!\n\n"
                message += f"Додано записів: {stats['added']}\n"
                if stats['errors'] > 0:
                    message += f"Помилок: {stats['errors']}\n"
                message += f"\n✓ Періоди автоматично розраховані для {stats['added']} осіб"

                QMessageBox.information(
                    self,
                    "Імпорт завершено",
                    message
                )

                self.accept()

            else:
                # СТАРИЙ ШЛЯХ: Імпорт в Excel
                progress.setLabelText("Імпорт в Excel файл...")
                success_count = 0
                error_count = 0

                target_sheet = self.excel_reader.workbook["Data"]
                next_row = target_sheet.max_row + 1

                for i, record in enumerate(all_records):
                    progress.setValue(30 + int(i / len(all_records) * 60))

                    try:
                        name = record["name"]

                        # Додаємо рядок
                        target_sheet.cell(row=next_row, column=1, value=self.month)  # A - місяць
                        target_sheet.cell(row=next_row, column=4, value=record["rank"])  # D - звання
                        target_sheet.cell(row=next_row, column=5, value=name)  # E - ПІБ
                        target_sheet.cell(row=next_row, column=7, value=record["position"])  # G - посада
                        target_sheet.cell(row=next_row, column=8, value=record["start_100"])  # H - початок 100
                        target_sheet.cell(row=next_row, column=9, value=record["end_100"])  # I - кінець 100
                        target_sheet.cell(row=next_row, column=10, value=record["start_30"])  # J - початок 30
                        target_sheet.cell(row=next_row, column=11, value=record["end_30"])  # K - кінець 30
                        target_sheet.cell(row=next_row, column=12, value=record["start_non"])  # L - початок не залучення
                        target_sheet.cell(row=next_row, column=13, value=record["end_non"])  # M - кінець не залучення

                        next_row += 1
                        success_count += 1
                    except Exception as e:
                        print(f"Помилка для {name}: {str(e)}")
                        error_count += 1

                progress.setValue(90)

                # Зберігаємо
                if self.excel_reader.save():
                    message = f"Дані успішно імпортовано!\n\n"
                    message += f"Додано записів: {success_count}\n"
                    if error_count > 0:
                        message += f"Помилок: {error_count}\n"
                    message += f"\n⚠️ Не забудь вказати статуси під час не залучення!"

                    QMessageBox.information(self, "Успіх", message)
                    progress.setValue(100)
                    self.accept()
                else:
                    QMessageBox.critical(self, "Помилка", "Не вдалось зберегти дані!")

        except Exception as e:
            progress.close()
            QMessageBox.critical(self, "Помилка", f"Помилка при імпорті:\n{str(e)}")

    def _update_periods_from_excel(self):
        """
        Оновлює періоди з основного Excel файлу після імпорту
        """
        try:
            from core.excel_reader import ExcelReader
            from core.data_processor import DataProcessor

            # Знаходимо основний Excel файл
            base_dir = get_base_dir()
            excel_path = os.path.join(base_dir, "D0A02800.xlsx")

            if not os.path.exists(excel_path):
                QMessageBox.warning(
                    self,
                    "Файл не знайдено",
                    f"Не знайдено файл {excel_path}\n\n"
                    f"Запустіть оновлення вручну через update_periods.bat"
                )
                return

            # Показуємо прогрес
            progress = QProgressDialog("Оновлення періодів з Excel...", None, 0, 100, self)
            progress.setWindowModality(Qt.WindowModal)
            progress.show()

            from PySide6.QtWidgets import QApplication
            QApplication.processEvents()

            # Завантажуємо Excel
            progress.setLabelText("Завантаження Excel файлу...")
            progress.setValue(10)
            QApplication.processEvents()

            excel_reader = ExcelReader(excel_path)
            excel_reader.load_workbook()

            progress.setLabelText("Оновлення періодів...")
            progress.setValue(30)
            QApplication.processEvents()

            # Отримуємо всіх servicemembers
            all_members = self.db_manager.get_all_servicemembers()
            total = len(all_members)

            updated_count = 0
            for i, member in enumerate(all_members):
                if i % 50 == 0:
                    progress.setValue(30 + int(i / total * 60))
                    QApplication.processEvents()

                name = member["name"]
                member_id = member["id"]

                # Отримати періоди з Excel
                periods_100_data = excel_reader.get_servicemember_data(name, "Періоди на 100")
                periods_30_data = excel_reader.get_servicemember_data(name, "Періоди на 30")

                # Збір та парсинг періодів
                periods_100 = []
                periods_30 = []

                for row in periods_100_data:
                    if row.get("periods"):
                        parsed = DataProcessor.parse_periods(row["periods"])
                        periods_100.extend(parsed)

                for row in periods_30_data:
                    if row.get("periods"):
                        parsed = DataProcessor.parse_periods(row["periods"])
                        periods_30.extend(parsed)

                if not periods_100 and not periods_30:
                    continue

                # Злиття та форматування
                merged_100 = DataProcessor.merge_consecutive_periods(periods_100)
                merged_30 = DataProcessor.merge_consecutive_periods(periods_30)
                formatted_100 = DataProcessor.format_periods_for_document(merged_100)
                formatted_30 = DataProcessor.format_periods_for_document(merged_30)

                # Оновити БД
                cursor = self.db_manager.connection.cursor()
                cursor.execute("DELETE FROM periods WHERE servicemember_id = ?", (member_id,))
                cursor.execute("DELETE FROM parsed_periods WHERE servicemember_id = ?", (member_id,))

                if formatted_100:
                    cursor.execute("""
                        INSERT INTO periods (servicemember_id, period_type, period_text)
                        VALUES (?, '100', ?)
                    """, (member_id, formatted_100))

                if formatted_30:
                    cursor.execute("""
                        INSERT INTO periods (servicemember_id, period_type, period_text)
                        VALUES (?, '30', ?)
                    """, (member_id, formatted_30))

                for start, end in merged_100:
                    cursor.execute("""
                        INSERT INTO parsed_periods (servicemember_id, period_type, start_date, end_date)
                        VALUES (?, '100', ?, ?)
                    """, (member_id, start.isoformat(), end.isoformat()))

                for start, end in merged_30:
                    cursor.execute("""
                        INSERT INTO parsed_periods (servicemember_id, period_type, start_date, end_date)
                        VALUES (?, '30', ?, ?)
                    """, (member_id, start.isoformat(), end.isoformat()))

                self.db_manager.connection.commit()
                updated_count += 1

            # ОНОВЛЕННЯ SERVICE_RECORDS ДЛЯ ІМПОРТОВАНОГО МІСЯЦЯ
            # Після оновлення parsed_periods потрібно також оновити service_records
            # для імпортованого місяця з датами start_30/end_30
            progress.setLabelText("Оновлення service_records з датами 30%...")
            progress.setValue(95)
            QApplication.processEvents()

            cursor = self.db_manager.connection.cursor()

            # Отримуємо всі service_records для імпортованого місяця
            cursor.execute("""
                SELECT sr.id, sr.servicemember_id
                FROM service_records sr
                WHERE sr.month = ?
            """, (self.month,))

            month_records = cursor.fetchall()

            for record_id, member_id in month_records:
                # Знаходимо parsed_periods типу '30' для цього servicemember
                # які перетинаються з імпортованим місяцем
                cursor.execute("""
                    SELECT start_date, end_date
                    FROM parsed_periods
                    WHERE servicemember_id = ? AND period_type = '30'
                    ORDER BY start_date
                """, (member_id,))

                periods_30 = cursor.fetchall()

                if not periods_30:
                    continue

                # Парсимо місяць (формат YYYY-MM)
                from datetime import datetime
                year, month = map(int, self.month.split('-'))

                # Знаходимо перший та останній періоди, що потрапляють в цей місяць
                start_30 = None
                end_30 = None

                for start_str, end_str in periods_30:
                    start = datetime.fromisoformat(start_str)
                    end = datetime.fromisoformat(end_str)

                    # Перевіряємо, чи період потрапляє в імпортований місяць
                    if start.year == year and start.month == month:
                        if start_30 is None:
                            start_30 = start
                    if end.year == year and end.month == month:
                        end_30 = end

                # Оновлюємо service_record
                if start_30 and end_30:
                    cursor.execute("""
                        UPDATE service_records
                        SET start_30 = ?, end_30 = ?
                        WHERE id = ?
                    """, (start_30.strftime("%d.%m.%Y"), end_30.strftime("%d.%m.%Y"), record_id))

            self.db_manager.connection.commit()

            progress.setValue(100)
            progress.close()

            QMessageBox.information(
                self,
                "Успіх",
                f"Періоди оновлено для {updated_count} військовослужбовців!"
            )

        except Exception as e:
            QMessageBox.critical(
                self,
                "Помилка",
                f"Помилка при оновленні періодів:\n{str(e)}\n\n"
                f"Запустіть оновлення вручну через update_periods.bat"
            )

    def format_date(self, date_value):
        """
        Форматує дату в DD.MM.YYYY
        """
        if isinstance(date_value, datetime):
            return date_value.strftime("%d.%m.%Y")
        elif isinstance(date_value, str):
            return date_value
        else:
            return str(date_value) if date_value else None

    def close_all_workbooks(self):
        """Закриває всі відкриті workbooks щоб звільнити файли"""
        for step_key in ["100", "30", "non"]:
            workbook = self.step_data[step_key].get("workbook")
            if workbook:
                try:
                    workbook.close()
                except:
                    pass
                self.step_data[step_key]["workbook"] = None

    def reject(self):
        """Закриття діалогу - звільняємо файли"""
        self.close_all_workbooks()
        super().reject()

    def accept(self):
        """Успішне закриття діалогу - звільняємо файли"""
        self.close_all_workbooks()
        super().accept()
