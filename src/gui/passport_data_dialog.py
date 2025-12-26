"""
Діалог для вибору джерела паспортних даних
"""
from PySide6.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QLabel,
    QLineEdit, QPushButton, QComboBox, QMessageBox,
    QFileDialog, QGroupBox, QRadioButton, QButtonGroup
)
from PySide6.QtCore import Qt
from openpyxl import load_workbook
from typing import Dict, Optional


class PassportDataDialog(QDialog):
    """
    Діалог для вибору способу введення паспортних даних:
    1. Ввести вручну (один раз для всіх)
    2. Завантажити з файлу Excel
    3. Пропустити (без паспортних даних)
    """

    def __init__(self, count: int, parent=None):
        """
        Args:
            count: Кількість рапортів для генерації
        """
        super().__init__(parent)
        self.count = count
        self.passport_data = {}  # {ПІБ: {"СЕРІЯ": ..., "НОМЕР": ...}}
        self.manual_data = {}  # {"СЕРІЯ": ..., "НОМЕР": ...} для всіх
        self.mode = "skip"  # "manual", "file", "skip"
        self.workbook = None  # Завантажений Excel файл

        self.init_ui()

    def init_ui(self):
        self.setWindowTitle("Паспортні дані")
        self.setMinimumWidth(500)

        layout = QVBoxLayout()

        # Стилі
        combo_style = "QComboBox { color: black; background-color: white; }"
        input_style = "QLineEdit { color: black; background-color: white; }"

        # Заголовок
        title = QLabel(f"Генерація {self.count} рапортів")
        title.setStyleSheet("font-size: 16px; font-weight: bold;")
        layout.addWidget(title)

        info = QLabel("Оберіть спосіб заповнення паспортних даних:")
        info.setStyleSheet("color: #666;")
        layout.addWidget(info)

        layout.addSpacing(15)

        # Радіо-кнопки вибору режиму
        self.radio_skip = QRadioButton("Пропустити паспортні дані")
        self.radio_manual = QRadioButton("Ввести вручну (однакові для всіх)")
        self.radio_file = QRadioButton("Завантажити з Excel файлу")

        self.radio_skip.setChecked(True)
        self.radio_skip.toggled.connect(self.on_mode_changed)
        self.radio_manual.toggled.connect(self.on_mode_changed)
        self.radio_file.toggled.connect(self.on_mode_changed)

        layout.addWidget(self.radio_skip)
        layout.addWidget(self.radio_manual)
        layout.addWidget(self.radio_file)

        layout.addSpacing(15)

        # Група для ручного введення
        self.manual_group = QGroupBox("Паспортні дані")
        manual_layout = QVBoxLayout()

        series_layout = QHBoxLayout()
        series_layout.addWidget(QLabel("Серія:"))
        self.series_input = QLineEdit()
        self.series_input.setPlaceholderText("АА (необов'язково)")
        self.series_input.setStyleSheet(input_style)
        series_layout.addWidget(self.series_input)
        manual_layout.addLayout(series_layout)

        number_layout = QHBoxLayout()
        number_layout.addWidget(QLabel("Номер:"))
        self.number_input = QLineEdit()
        self.number_input.setPlaceholderText("123456")
        self.number_input.setStyleSheet(input_style)
        number_layout.addWidget(self.number_input)
        manual_layout.addLayout(number_layout)

        self.manual_group.setLayout(manual_layout)
        self.manual_group.setVisible(False)
        layout.addWidget(self.manual_group)

        # Група для файлу
        self.file_group = QGroupBox("Файл з паспортними даними")
        file_layout = QVBoxLayout()

        file_info = QLabel("Excel файл повинен містити колонки: ПІБ, Серія, Номер")
        file_info.setStyleSheet("color: #666; font-style: italic;")
        file_layout.addWidget(file_info)

        file_select_layout = QHBoxLayout()
        self.file_label = QLabel("Файл не обрано")
        self.file_label.setStyleSheet("color: gray;")
        file_select_layout.addWidget(self.file_label)

        file_btn = QPushButton("Обрати файл...")
        file_btn.clicked.connect(self.select_file)
        file_select_layout.addWidget(file_btn)
        file_layout.addLayout(file_select_layout)

        # Вибір аркуша
        sheet_layout = QHBoxLayout()
        sheet_layout.addWidget(QLabel("Аркуш:"))
        self.sheet_combo = QComboBox()
        self.sheet_combo.setStyleSheet(combo_style)
        self.sheet_combo.setEnabled(False)
        self.sheet_combo.currentTextChanged.connect(self.on_sheet_changed)
        sheet_layout.addWidget(self.sheet_combo)
        sheet_layout.addStretch()
        file_layout.addLayout(sheet_layout)

        # Налаштування колонок (випадаючі списки з заголовками)
        columns_layout = QVBoxLayout()

        # Перший рядок колонок
        row1 = QHBoxLayout()
        row1.addWidget(QLabel("Колонка ПІБ:"))
        self.name_col_combo = QComboBox()
        self.name_col_combo.setStyleSheet(combo_style)
        self.name_col_combo.setMinimumWidth(150)
        row1.addWidget(self.name_col_combo)
        row1.addStretch()
        columns_layout.addLayout(row1)

        # Другий рядок колонок
        row2 = QHBoxLayout()
        row2.addWidget(QLabel("Серія:"))
        self.series_col_combo = QComboBox()
        self.series_col_combo.setStyleSheet(combo_style)
        self.series_col_combo.setMinimumWidth(150)
        row2.addWidget(self.series_col_combo)

        row2.addWidget(QLabel("Номер:"))
        self.number_col_combo = QComboBox()
        self.number_col_combo.setStyleSheet(combo_style)
        self.number_col_combo.setMinimumWidth(150)
        row2.addWidget(self.number_col_combo)
        row2.addStretch()
        columns_layout.addLayout(row2)

        file_layout.addLayout(columns_layout)

        # Статус завантаження
        self.load_status = QLabel("")
        file_layout.addWidget(self.load_status)

        self.file_group.setLayout(file_layout)
        self.file_group.setVisible(False)
        layout.addWidget(self.file_group)

        layout.addStretch()

        # Кнопки
        button_layout = QHBoxLayout()

        ok_btn = QPushButton("Продовжити")
        ok_btn.setStyleSheet("background-color: #4caf50; color: white; font-weight: bold; padding: 8px 16px;")
        ok_btn.clicked.connect(self.on_accept)
        button_layout.addWidget(ok_btn)

        cancel_btn = QPushButton("Скасувати")
        cancel_btn.clicked.connect(self.reject)
        button_layout.addWidget(cancel_btn)

        layout.addLayout(button_layout)
        self.setLayout(layout)

    def on_mode_changed(self):
        """Обробник зміни режиму"""
        self.manual_group.setVisible(self.radio_manual.isChecked())
        self.file_group.setVisible(self.radio_file.isChecked())

        if self.radio_skip.isChecked():
            self.mode = "skip"
        elif self.radio_manual.isChecked():
            self.mode = "manual"
        elif self.radio_file.isChecked():
            self.mode = "file"

    def column_letter_to_index(self, letter: str) -> int:
        """Конвертує букву стовпця в індекс (A=0, B=1, ...)"""
        letter = letter.upper().strip()
        result = 0
        for char in letter:
            result = result * 26 + (ord(char) - ord('A') + 1)
        return result - 1  # 0-based

    def select_file(self):
        """Вибір файлу з паспортними даними"""
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Оберіть файл з паспортними даними",
            "",
            "Excel Files (*.xlsx *.xls)"
        )

        if file_path:
            try:
                self.workbook = load_workbook(file_path, data_only=True)
                self.file_label.setText(file_path)
                self.file_label.setStyleSheet("color: green;")

                # Заповнюємо список аркушів
                self.sheet_combo.clear()
                self.sheet_combo.addItems(self.workbook.sheetnames)
                self.sheet_combo.setEnabled(True)

                # Завантажуємо дані з першого аркуша
                self.load_passport_data()

            except Exception as e:
                self.load_status.setText(f"Помилка: {str(e)}")
                self.load_status.setStyleSheet("color: red;")
                self.workbook = None

    def on_sheet_changed(self):
        """Обробник зміни аркуша - заповнює список колонок"""
        if not self.workbook:
            return

        sheet_name = self.sheet_combo.currentText()
        if not sheet_name:
            return

        try:
            sheet = self.workbook[sheet_name]

            # Отримуємо заголовки з першого рядка
            headers = []
            for col_idx, cell in enumerate(next(sheet.iter_rows(min_row=1, max_row=1, values_only=True)), 1):
                col_letter = self.index_to_column_letter(col_idx)
                header_text = str(cell) if cell else f"(Колонка {col_letter})"
                headers.append((col_idx, f"{col_letter}: {header_text}"))

            # Заповнюємо комбобокси
            self.name_col_combo.clear()
            self.series_col_combo.clear()
            self.number_col_combo.clear()

            for col_idx, header in headers:
                self.name_col_combo.addItem(header, col_idx)
                self.series_col_combo.addItem(header, col_idx)
                self.number_col_combo.addItem(header, col_idx)

            # Автоматично вибираємо колонки якщо знайдено відповідні заголовки
            self.auto_select_columns(headers)

            # Завантажуємо дані
            self.load_passport_data()

        except Exception as e:
            self.load_status.setText(f"Помилка: {str(e)}")
            self.load_status.setStyleSheet("color: red;")

    def index_to_column_letter(self, col_idx: int) -> str:
        """Конвертує індекс колонки в букву (1=A, 2=B, ...)"""
        result = ""
        while col_idx > 0:
            col_idx, remainder = divmod(col_idx - 1, 26)
            result = chr(65 + remainder) + result
        return result

    def auto_select_columns(self, headers):
        """Автоматично вибирає колонки за назвами заголовків"""
        name_keywords = ["піб", "прізвище", "ім'я", "name", "пiб"]
        series_keywords = ["серія", "серия", "series"]
        number_keywords = ["номер", "number", "№"]

        for i, (col_idx, header) in enumerate(headers):
            header_lower = header.lower()

            # Шукаємо колонку ПІБ
            for keyword in name_keywords:
                if keyword in header_lower:
                    self.name_col_combo.setCurrentIndex(i)
                    break

            # Шукаємо колонку Серія
            for keyword in series_keywords:
                if keyword in header_lower:
                    self.series_col_combo.setCurrentIndex(i)
                    break

            # Шукаємо колонку Номер
            for keyword in number_keywords:
                if keyword in header_lower:
                    self.number_col_combo.setCurrentIndex(i)
                    break

    def load_passport_data(self):
        """Завантажує паспортні дані з обраного аркуша"""
        if not self.workbook:
            return

        try:
            sheet_name = self.sheet_combo.currentText()
            if not sheet_name:
                return

            sheet = self.workbook[sheet_name]

            # Отримуємо індекси колонок з комбобоксів (1-based)
            name_col = self.name_col_combo.currentData()
            series_col = self.series_col_combo.currentData()
            number_col = self.number_col_combo.currentData()

            if not name_col:
                self.load_status.setText("Оберіть колонку ПІБ")
                self.load_status.setStyleSheet("color: orange;")
                return

            # Конвертуємо в 0-based для доступу до row[]
            name_col = name_col - 1
            series_col = (series_col - 1) if series_col else None
            number_col = (number_col - 1) if number_col else None

            self.passport_data = {}
            count = 0

            # Визначаємо максимальний індекс колонки
            max_col = name_col
            if series_col is not None:
                max_col = max(max_col, series_col)
            if number_col is not None:
                max_col = max(max_col, number_col)

            for row in sheet.iter_rows(min_row=2, values_only=True):
                if len(row) <= max_col:
                    continue

                name = str(row[name_col]).strip() if row[name_col] else None
                if not name or name == "None":
                    continue

                series = ""
                number = ""

                if series_col is not None and series_col < len(row) and row[series_col]:
                    series = str(row[series_col]).strip().upper()
                    if series == "NONE":
                        series = ""

                if number_col is not None and number_col < len(row) and row[number_col]:
                    number = str(row[number_col]).strip()
                    if number == "NONE":
                        number = ""

                self.passport_data[name] = {
                    "СЕРІЯ": series,
                    "НОМЕР": number
                }
                count += 1

            self.load_status.setText(f"Завантажено {count} записів")
            self.load_status.setStyleSheet("color: green;")

        except Exception as e:
            self.load_status.setText(f"Помилка: {str(e)}")
            self.load_status.setStyleSheet("color: red;")
            self.passport_data = {}

    def on_accept(self):
        """Обробник кнопки OK"""
        if self.mode == "manual":
            self.manual_data = {
                "СЕРІЯ": self.series_input.text().strip().upper(),
                "НОМЕР": self.number_input.text().strip()
            }
        elif self.mode == "file":
            if not self.passport_data:
                QMessageBox.warning(self, "Помилка", "Спочатку завантажте файл з паспортними даними!")
                return

        # Закриваємо workbook щоб звільнити файл
        self.close_workbook()
        self.accept()

    def close_workbook(self):
        """Закриває workbook і звільняє файл"""
        if self.workbook:
            try:
                self.workbook.close()
            except:
                pass
            self.workbook = None

    def reject(self):
        """Закриття діалогу - звільняємо файл"""
        self.close_workbook()
        super().reject()

    def get_passport_for_name(self, name: str) -> Dict[str, str]:
        """
        Отримує паспортні дані для конкретного ПІБ

        Args:
            name: ПІБ військовослужбовця

        Returns:
            {"СЕРІЯ": ..., "НОМЕР": ...}
        """
        if self.mode == "skip":
            return {"СЕРІЯ": "", "НОМЕР": ""}
        elif self.mode == "manual":
            return self.manual_data
        elif self.mode == "file":
            # Шукаємо точне співпадіння
            if name in self.passport_data:
                return self.passport_data[name]

            # Шукаємо часткове співпадіння (без урахування регістру)
            name_upper = name.upper()
            for key, value in self.passport_data.items():
                if key.upper() == name_upper:
                    return value

            # Не знайдено - повертаємо порожнє
            return {"СЕРІЯ": "", "НОМЕР": ""}

        return {"СЕРІЯ": "", "НОМЕР": ""}
