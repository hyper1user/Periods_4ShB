"""
Валідація даних перед обробкою
"""
import os
from typing import Tuple
from openpyxl import load_workbook


def validate_excel_file(file_path: str) -> Tuple[bool, str]:
    """
    Валідує Excel файл

    Args:
        file_path: Шлях до Excel файлу

    Returns:
        (True, "") якщо файл валідний, (False, "повідомлення про помилку") інакше
    """
    # Перевірка існування файлу
    if not os.path.exists(file_path):
        return False, f"Файл не знайдено: {file_path}"

    # Перевірка розширення
    if not file_path.endswith(('.xlsx', '.xlsm')):
        return False, "Файл повинен мати розширення .xlsx або .xlsm"

    # Спроба відкрити файл
    try:
        wb = load_workbook(file_path, read_only=True, keep_vba=True)

        # Перевірка наявності обов'язкових аркушів
        required_sheets = ["Data", "Періоди на 100", "Періоди на 30"]
        missing_sheets = []

        for sheet_name in required_sheets:
            if sheet_name not in wb.sheetnames:
                missing_sheets.append(sheet_name)

        wb.close()

        if missing_sheets:
            return False, f"Відсутні обов'язкові аркуші: {', '.join(missing_sheets)}"

        return True, ""

    except Exception as e:
        return False, f"Помилка при читанні файлу: {str(e)}"


def validate_servicemember_data(data: dict) -> Tuple[bool, list[str]]:
    """
    Валідує дані військовослужбовця

    Args:
        data: Словник з даними військовослужбовця

    Returns:
        (True, []) якщо дані валідні, (False, [список помилок]) інакше
    """
    errors = []

    # Перевірка обов'язкових полів
    required_fields = ["name", "rank", "position"]

    for field in required_fields:
        if not data.get(field):
            errors.append(f"Відсутнє обов'язкове поле: {field}")

    # Перевірка наявності періодів
    if not data.get("periods"):
        errors.append("Відсутні періоди для цього військовослужбовця")

    return (len(errors) == 0, errors)


def validate_periods(periods_str: str) -> bool:
    """
    Валідує рядок з періодами

    Args:
        periods_str: Рядок з періодами

    Returns:
        True якщо рядок валідний, False інакше
    """
    if not periods_str or not isinstance(periods_str, str):
        return False

    # Перевірка наявності дат у форматі DD.MM.YYYY
    import re
    date_pattern = r'\d{2}\.\d{2}\.\d{4}'
    dates = re.findall(date_pattern, periods_str)

    # Має бути парна кількість дат (початок та кінець)
    return len(dates) >= 2 and len(dates) % 2 == 0
